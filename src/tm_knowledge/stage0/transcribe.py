"""The return leg: workbook in, validated records out (parallel track P8).

This is the mechanism by which "an agent will transcribe and validate" stops
being a promise in `STAGE-0-INPUT-GUIDE.md` §6 and becomes a command.

**Transcription may reshape, never supply.** A blank cell becomes a null and is
reported as a gap. If a relationship arrives without a `modality`, the record is
written without one and the gap is queued — it is never inferred from the
sentence's grammar, because whether a "may" is possibility or permission is a
legal reading (guide §5.4). Nothing in this module has a default value for a
judgement field, and adding one would be authoring content.

Three things it refuses to do, all of them for the same reason — writing into
`eval/gold/` is writing into approved space (CLAUDE.md rule 4):

- It will not write a record that does not validate. A malformed record in the
  gold set is a measurement standard that is wrong, and the harness would then
  certify everything against it.
- It will not write a row whose non-nullable required fields are blank. That row
  is not yet a record; saying so is more useful than writing a stub.
- It will not write at all without `--write`. The default is a dry run that
  prints what would change.

Re-running over unchanged input rewrites nothing: the YAML is rendered
deterministically and compared before writing, so an unchanged workbook produces
an unchanged git status.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml

from tm_knowledge.stage0 import goldset
from tm_knowledge.stage0.intake import Column, Sheet, sheet_for, sheets
from tm_knowledge.stage0.schemas import property_order, required_fields, validate

__all__ = ["Transcription", "read_workbook", "write_records"]


class WorkbookMismatch(Exception):
    """The workbook's columns are not the ones the schemas describe."""


@dataclass(frozen=True, slots=True)
class Problem:
    """A row that could not become a record, said where a person can find it."""

    sheet: str
    row: int | None
    message: str

    def __str__(self) -> str:
        where = f"{self.sheet}" + (f" row {self.row}" if self.row else "")
        return f"{where}: {self.message}"


@dataclass
class Transcription:
    """What came out of one workbook."""

    records: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    problems: list[Problem] = field(default_factory=list)
    #: (record type, id, field) for every judgement field that arrived blank.
    blanks: list[tuple[str, str, str]] = field(default_factory=list)
    #: Record types whose sheet held no rows at all.
    empty_sheets: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return sum(len(records) for records in self.records.values())

    def summary(self) -> str:
        return (
            f"{self.total} record(s) from {len(self.records)} sheet(s); "
            f"{len(self.problems)} rejected row(s); {len(self.blanks)} blank "
            "judgement field(s)"
        )


# ---------------------------------------------------------------------------
# Reading cells
# ---------------------------------------------------------------------------


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell(column: Column, value: Any, sheet: str, row: int) -> Any:
    """One cell to one Python value. Raises on a value the schema cannot hold."""
    if column.kind == "list":
        raw = _text(value)
        if raw is None:
            return []
        return [line.strip() for line in raw.splitlines() if line.strip()]

    text = _text(value)
    if text is None:
        return None

    if column.kind == "boolean":
        lowered = text.lower()
        if lowered in ("true", "yes", "y", "1"):
            return True
        if lowered in ("false", "no", "n", "0"):
            return False
        raise ValueError(f"{column.header} is {text!r}, which is not true or false")

    if column.kind == "number":
        try:
            return int(float(text))
        except ValueError as error:
            raise ValueError(f"{column.header} is {text!r}, which is not a number") from error

    if column.enum:
        for allowed in column.enum:
            if text == str(allowed):
                return allowed
        raise ValueError(
            f"{column.header} is {text!r}; it takes one of "
            + ", ".join(str(v) for v in column.enum)
            + ". Leave it blank rather than choosing the nearest one"
        )

    return text


def _prune_optional(record: dict[str, Any], record_type: str) -> None:
    """Drop optional keys that arrived empty; keep every required one.

    The asymmetry is ADR-0027's, and it is what keeps a gap visible. A required
    key is written even when it is null, because that null *is* the gap the
    coverage report names. An optional key that nobody filled in is not a gap —
    writing `notes: null` on every record would bury the real ones in noise, and
    it would make the transcriber's output differ from an expert's own file for
    no reason.
    """
    required = required_fields(record_type)
    for key in [k for k in record if k not in required]:
        if record[key] is None or record[key] == [] or record[key] == "":
            del record[key]


def _reorder(record: dict[str, Any], record_type: str) -> None:
    """Put the keys back into schema order, in place.

    A child sheet attaches its field after the parent row was built, so
    `relevant` and `expected_inferences` would otherwise land at the end of the
    record — correct YAML, and a diff that reads nothing like the template.
    """
    ordered = {
        key: record[key] for key in property_order(record_type) if key in record
    }
    record.clear()
    record.update(ordered)


def _nest(record: dict[str, Any], path: tuple[str, ...], value: Any) -> None:
    node = record
    for key in path[:-1]:
        node = node.setdefault(key, {})
    node[path[-1]] = value


def _collapse_spans(record: dict[str, Any]) -> None:
    """`span.start` / `span.end` back into the `[start, end]` the schema wants."""
    span = record.get("span")
    if not isinstance(span, dict):
        return
    start, end = span.get("start"), span.get("end")
    if start is None and end is None:
        record["span"] = None
        return
    if start is None or end is None:
        raise ValueError("span.start and span.end must be given together, or neither")
    record["span"] = [start, end]


# ---------------------------------------------------------------------------
# Reading sheets
# ---------------------------------------------------------------------------


def _headers(worksheet, spec: Sheet) -> dict[str, int]:
    """Header name -> column index, checked against the layout in both directions."""
    found: dict[str, int] = {}
    for index, cell in enumerate(worksheet[1], start=1):
        name = _text(cell.value)
        if name is None:
            continue
        if name in found:
            raise WorkbookMismatch(f"{spec.name}: the column {name!r} appears twice")
        found[name] = index

    expected = {column.header for column in spec.columns}
    missing = expected - set(found)
    unknown = set(found) - expected
    if missing or unknown:
        raise WorkbookMismatch(
            f"{spec.name}: the sheet does not match the schemas. "
            + (f"Missing: {', '.join(sorted(missing))}. " if missing else "")
            + (f"Unknown: {', '.join(sorted(unknown))}. " if unknown else "")
            + "Regenerate the workbook with `tmk-workbook` rather than adding "
            "columns by hand — a column the transcriber does not know is a field "
            "nobody collects."
        )
    return found


def _rows(worksheet, spec: Sheet, result: Transcription):
    """(row number, record) for every row that became a record."""
    positions = _headers(worksheet, spec)
    for number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        values = {header: row[index - 1].value for header, index in positions.items()}
        if all(_text(value) is None for value in values.values()):
            continue

        record: dict[str, Any] = {}
        failed = False
        for column in spec.columns:
            try:
                _nest(record, column.path, _cell(column, values[column.header], spec.name, number))
            except ValueError as error:
                result.problems.append(Problem(spec.name, number, str(error)))
                failed = True
        if failed:
            continue
        try:
            _collapse_spans(record)
        except ValueError as error:
            result.problems.append(Problem(spec.name, number, str(error)))
            continue
        if not spec.is_child:
            _prune_optional(record, spec.record_type)

        blank_required = [
            column.header
            for column in spec.columns
            if column.required and not column.nullable
            and record.get(column.path[0]) in (None, "", [])
        ]
        if blank_required:
            result.problems.append(
                Problem(
                    spec.name,
                    number,
                    f"{', '.join(blank_required)} is blank, so the row is not yet a "
                    "record. Nothing here will fill it in",
                )
            )
            continue
        yield number, record


def read_workbook(path: Path) -> Transcription:
    """Read a filled workbook into records. Writes nothing."""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:  # pragma: no cover - depends on install
        raise ModuleNotFoundError(
            "the intake path needs openpyxl: pip install -e '.[intake]'"
        ) from error

    workbook = load_workbook(path, data_only=True)
    result = Transcription()

    parents: dict[str, dict[str, dict[str, Any]]] = {}
    for spec in sheets():
        if spec.is_child or spec.name not in workbook.sheetnames:
            continue
        rows = dict(_rows(workbook[spec.name], spec, result))
        if not rows:
            result.empty_sheets.append(spec.record_type)
            continue
        by_id: dict[str, dict[str, Any]] = {}
        for number, record in rows.items():
            identifier = record.get("id")
            if identifier in by_id:
                result.problems.append(
                    Problem(spec.name, number, f"id {identifier} is used twice")
                )
                continue
            by_id[identifier] = record
        parents[spec.record_type] = by_id

    # Child sheets fill a repeating field on a parent that must already exist.
    for spec in sheets():
        if not spec.is_child or spec.name not in workbook.sheetnames:
            continue
        by_id = parents.get(spec.record_type, {})
        for parent in by_id.values():
            parent.setdefault(spec.parent_field, [])
        for number, entry in _rows(workbook[spec.name], spec, result):
            parent_id = entry.pop("parent_id")
            parent = by_id.get(parent_id)
            if parent is None:
                result.problems.append(
                    Problem(
                        spec.name,
                        number,
                        f"parent_id {parent_id} names no record on the "
                        f"`{_parent_sheet(spec)}` sheet",
                    )
                )
                continue
            parent[spec.parent_field].append(entry)

    for record_type, by_id in parents.items():
        kept: list[dict[str, Any]] = []
        for identifier, record in by_id.items():
            _reorder(record, record_type)
            errors = validate(record, record_type)
            if errors:
                for error in errors:
                    result.problems.append(
                        Problem(_sheet_name(record_type), None, f"{identifier}: {error}")
                    )
                continue
            kept.append(record)
            result.blanks.extend(
                (record_type, identifier, name) for name in _blank_fields(record)
            )
        if kept:
            result.records[record_type] = kept
    return result


def _sheet_name(record_type: str) -> str:
    return next(s.name for s in sheets() if s.record_type == record_type and not s.is_child)


def _parent_sheet(spec: Sheet) -> str:
    return _sheet_name(spec.record_type)


def _blank_fields(record: dict[str, Any], prefix: str = "") -> list[str]:
    """Every key that arrived null or empty. What P10 turns into a worklist."""
    blank: list[str] = []
    for key, value in record.items():
        name = f"{prefix}{key}"
        if value is None or value == [] or value == "":
            blank.append(name)
        elif isinstance(value, dict):
            blank.extend(_blank_fields(value, f"{name}."))
    return blank


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------


def render(records: list[dict[str, Any]]) -> str:
    """One gold file's YAML, deterministically.

    Key order is the schema's order, because that is the order the workbook's
    columns are in and the order the guide explains the fields in. `sort_keys`
    would reorder every record into alphabetical soup and make the first diff
    after this change unreadable.
    """
    return yaml.safe_dump(
        records,
        sort_keys=False,
        allow_unicode=True,
        default_flow_style=False,
        width=100,
    )


def write_records(
    transcription: Transcription,
    gold_dir: Path | None = None,
    *,
    write: bool = False,
) -> list[tuple[Path, str]]:
    """Write each record type's file. Returns (path, outcome) per file.

    Outcomes: `written`, `unchanged`, `would write` (a dry run). A record type
    whose sheet held no rows is not touched at all — an empty sheet means "I have
    nothing for this yet", never "delete what is there".
    """
    gold_dir = gold_dir or goldset.GOLD_DIR
    outcomes: list[tuple[Path, str]] = []
    for record_type, records in sorted(transcription.records.items()):
        path = gold_dir / goldset.FILE_FOR[record_type]
        text = render(records)
        current = path.read_text(encoding="utf-8") if path.exists() else None
        if current == text:
            outcomes.append((path, "unchanged"))
            continue
        if not write:
            outcomes.append((path, "would write"))
            continue
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text, encoding="utf-8")
        outcomes.append((path, "written"))
    return outcomes
