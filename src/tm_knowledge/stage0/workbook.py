"""Generate the intake workbook (parallel track P7).

`STAGE-0-INPUT-GUIDE.md` §6 tells the owner not to write YAML. This is what makes
that true: one sheet per record type, the guide's column order, enum cells as
dropdowns, and the shape rules as notes on the header cells.

**No example rows.** Not one, not even a clearly-marked one. The guide's
`«placeholder»` convention exists because a plausible filled row anchors the
reviewer and gets copied forward, and a workbook is the format where copying a
row is a keystroke (CLAUDE.md rule 1, parallel track §4 P7). The guidance lives
in cell notes and on the `how to use this` sheet, where it cannot be mistaken for
data.

The workbook regenerates from `eval/schemas/` through `intake.py`, so a schema
change cannot leave it stale — and `tmk-transcribe` reads the same layout back,
so the round trip preserves every field.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from tm_knowledge.stage0.intake import Column, Sheet, sheets
from tm_knowledge.stage0.schemas import RECORD_TYPES

__all__ = ["build", "write", "sheets"]

#: Where the enum lists live. Hidden, because it is machinery: a person who
#: unhides it and edits it breaks the dropdowns rather than changing the schema.
ENUM_SHEET = "_enums"

GUIDE_SHEET = "how to use this"

_HEADER_FILL = "FFEFEFEF"
_REQUIRED_FILL = "FFDCE9F5"


def _guidance() -> list[tuple[str, str]]:
    """The rules that govern filling this in. Shape rules only — no content."""
    return [
        (
            "One sheet per record type",
            "Fill in whichever sheets you have content for. An empty sheet is fine "
            "and is reported as a gap, not as an error.",
        ),
        (
            "One row per record",
            "The `id` column is the record's identifier. Allocate them in order "
            "and never reuse one, including after deleting a row — a withdrawn id "
            "goes in eval/gold/retired-ids.yaml.",
        ),
        (
            "Lists go one value per line",
            "In a list cell (refs, labels, ids), press Alt+Enter between values. "
            "Do not use commas or semicolons: refs contain punctuation, and a "
            "separator that can appear inside a value loses data silently.",
        ),
        (
            "The GS-- and GX-- sheets are continuations",
            "A search question's graded passages and a reasoning expectation's "
            "inferences each get their own sheet, one row per entry, linked by "
            "`parent_id`. Repeat the parent id on every row that belongs to it.",
        ),
        (
            "Leave a judgement blank rather than guessing",
            "A blank cell is transcribed as a gap and reported. It is never filled "
            "in for you, and nothing in this project may infer a tier, a modality "
            "or a type from the sentence's grammar — that is a legal reading.",
        ),
        (
            "approved_by and approved_date are the record's authority",
            "Without them the record is transcribed but not approved, and the "
            "harness reports it as awaiting approval. They are what distinguishes "
            "this content from a machine's candidate.",
        ),
        (
            "Blue headers are required by the schema",
            "Required means the field must be present, not that you must have an "
            "answer. Blank is a legitimate value for every judgement field.",
        ),
        (
            "Spans are character offsets into the passage text",
            "Take them from the Pass B worksheet (`tmk-worksheet`), which prints "
            "every ref, heading path and content hash already. You should never "
            "have to type a ref or a hash by hand.",
        ),
        (
            "Nothing here is an example",
            "Every sheet ships empty on purpose. A filled example row would be "
            "copied, and a plausible answer that nobody approved is worse than a "
            "blank one.",
        ),
    ]


def _enum_ranges(workbook) -> dict[str, str]:
    """Write every enum onto the hidden sheet, and return its range per column."""
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(ENUM_SHEET)
    ranges: dict[str, str] = {}
    index = 0
    for spec in sheets():
        for column in spec.columns:
            if not column.enum:
                continue
            key = f"{spec.name}!{column.header}"
            index += 1
            letter = get_column_letter(index)
            sheet.cell(row=1, column=index, value=key)
            for offset, value in enumerate(column.enum, start=2):
                sheet.cell(row=offset, column=index, value=value)
            ranges[key] = (
                f"'{ENUM_SHEET}'!${letter}$2:${letter}${len(column.enum) + 1}"
            )
    sheet.sheet_state = "hidden"
    return ranges


def _write_sheet(workbook, spec: Sheet, ranges: dict[str, str]) -> None:
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    sheet = workbook.create_sheet(spec.name)
    for index, column in enumerate(spec.columns, start=1):
        cell = sheet.cell(row=1, column=index, value=column.header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid", fgColor=_REQUIRED_FILL if column.required else _HEADER_FILL
        )
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        note = _note(column)
        if note:
            cell.comment = Comment(note, "tmk-workbook", height=180, width=340)

        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = _width(column)

        if column.enum:
            validation = DataValidation(
                type="list",
                formula1=ranges[f"{spec.name}!{column.header}"],
                allow_blank=True,
                showDropDown=False,
            )
            validation.error = (
                f"{column.header} takes one of: "
                + ", ".join(str(value) for value in column.enum)
                + ". Leave it blank rather than inventing a value."
            )
            validation.errorTitle = "Not one of the allowed values"
            sheet.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}1000")
        elif column.kind == "boolean":
            validation = DataValidation(
                type="list", formula1='"TRUE,FALSE"', allow_blank=True
            )
            sheet.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}1000")

    sheet.freeze_panes = "A2"


def _note(column: Column) -> str:
    parts = []
    if column.kind == "list":
        parts.append("A list — one value per line (Alt+Enter).")
    if column.enum:
        parts.append("One of: " + ", ".join(str(value) for value in column.enum) + ".")
    if column.required and column.nullable:
        parts.append("Required by the schema, and may be left blank: a blank is "
                     "reported as a gap, never filled in.")
    elif column.required:
        parts.append("Required. A row without it is not yet a record and will not "
                     "be transcribed.")
    if column.note:
        parts.append(column.note)
    return "\n\n".join(parts)


def _width(column: Column) -> int:
    if column.kind in ("number", "boolean"):
        return 14
    if column.kind == "list":
        return 34
    if column.enum:
        return 24
    return 40 if column.header in ("question", "query", "prohibited", "why") else 28


def _write_guide(workbook) -> None:
    from openpyxl.styles import Alignment, Font

    sheet = workbook.create_sheet(GUIDE_SHEET, 0)
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 96

    sheet["A1"] = "Stage 0 intake workbook"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = (
        "Generated from eval/schemas/ by `tmk-workbook`. Regenerate it rather than "
        "editing the columns — the transcriber reads the same layout, and a column "
        "added by hand is a field nobody collects."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A2:B2")
    sheet.row_dimensions[2].height = 32

    row = 4
    for heading, body in _guidance():
        sheet.cell(row=row, column=1, value=heading).font = Font(bold=True)
        cell = sheet.cell(row=row, column=2, value=body)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 44
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="When you are done").font = Font(bold=True)
    sheet.cell(
        row=row,
        column=2,
        value=(
            "Hand the file back. `tmk-transcribe` turns it into records in "
            "eval/gold/, validates them, and reports every blank judgement field "
            "as a gap. It reshapes; it never supplies."
        ),
    ).alignment = Alignment(wrap_text=True, vertical="top")


def build(*, generated: str | None = None):
    """Build the workbook in memory and return it."""
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as error:  # pragma: no cover - depends on install
        raise ModuleNotFoundError(
            "the intake path needs openpyxl: pip install -e '.[intake]'"
        ) from error

    workbook = Workbook()
    workbook.remove(workbook.active)
    ranges = _enum_ranges(workbook)
    for spec in sheets():
        _write_sheet(workbook, spec, ranges)
    _write_guide(workbook)
    workbook.properties.title = "TM-Knowledge — Stage 0 intake"
    workbook.properties.description = (
        "Generated by tmk-workbook on "
        f"{generated or datetime.now(timezone.utc).strftime('%Y-%m-%d')} from "
        f"eval/schemas/ ({len(RECORD_TYPES)} record types). Contains no example "
        "content by design."
    )
    return workbook


def write(path: Path, *, generated: str | None = None) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    build(generated=generated).save(path)
    return path


# ---------------------------------------------------------------------------
# Filling one back in
# ---------------------------------------------------------------------------


def fill(workbook, records: dict[str, list[dict]]) -> None:
    """Write records into an empty workbook, in the layout `transcribe` reads.

    It exists because the round trip has to be provable: P7's done-criterion is
    that a workbook survives a pass through P8 with every field intact, and that
    cannot be asserted without a way to put records in. It is deliberately **not**
    a command — generating an empty workbook and pre-filling one with content are
    different decisions, and only the first has been made.
    """
    from openpyxl.utils import get_column_letter

    del get_column_letter  # imported to fail early if openpyxl is a stub

    for spec in sheets():
        sheet = workbook[spec.name]
        row = 2
        for record in records.get(spec.record_type, ()):
            entries = [record] if not spec.is_child else [
                {"parent_id": record.get("id"), **entry}
                for entry in (record.get(spec.parent_field) or ())
            ]
            for entry in entries:
                for index, column in enumerate(spec.columns, start=1):
                    sheet.cell(row=row, column=index, value=_flatten(entry, column))
                row += 1


def _flatten(record: dict, column: Column):
    """One record's value for one column, in the cell encoding `intake` fixes."""
    node = record
    for key in column.path[:-1]:
        if not isinstance(node, dict):
            return None
        node = node.get(key)
        if node is None:
            return None
    if column.is_span_part:
        span = record.get("span")
        if not span:
            return None
        return span[0 if column.path[-1] == "start" else 1]
    if not isinstance(node, dict):
        return None
    value = node.get(column.path[-1])
    if value is None:
        return None
    if column.kind == "list":
        return "\n".join(str(item) for item in value) or None
    return value
