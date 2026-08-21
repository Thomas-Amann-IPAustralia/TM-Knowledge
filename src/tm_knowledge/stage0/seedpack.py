"""Render the seed set into the two things an expert can actually mark up.

`review/seed/*.seed.yaml` is the source of truth and nobody should have to read
it. This module turns it into:

- **the review pack** — Markdown, one section per record, with the passage the
  record rests on quoted underneath it and the span highlighted. It is designed
  to be read in order and scribbled on, on screen or on paper. Everything an
  expert needs to rule on a record is on the page, so nothing sends them back to
  the corpus;
- **the review workbook** — the intake workbook's own layout, pre-filled with
  the seed records, plus three review columns at the right-hand end. Correct a
  cell, set the verdict, put your name in `approved_by`, and `tmk-transcribe`
  reads it straight into `eval/gold/`. That is the whole point: the correction
  path and the authoring path are one path, so an expert who prefers to type
  over a draft never leaves the workflow the intake guide already describes.

**The intake workbook stays empty.** HANDOFF §4 forbids an example row in
`stage0-intake.xlsx` and that has not changed — a blank form is still the right
thing for someone composing from scratch, and a filled row in it would be copied
without thought. This is a *different file*, with a different name, a different
first sheet and a verdict column on every row (ADR-0044).
"""

from __future__ import annotations

import textwrap
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from tm_knowledge.stage0 import workbook as workbook_module
from tm_knowledge.stage0.intake import REVIEW_COLUMNS, sheets
from tm_knowledge.stage0.schemas import RECORD_TYPES
from tm_knowledge.stage0.seed import Resolution, SeedSet, VERDICTS
from tm_knowledge.upstream.loader import Corpus

__all__ = ["render_pack", "build_workbook", "write_workbook", "PACK_HEADINGS"]

#: Record type -> what to call it on the page, and the guide section that
#: explains it. The section numbers are `eval/STAGE-0-INPUT-GUIDE.md`'s.
PACK_HEADINGS: dict[str, tuple[str, str]] = {
    "competency_question": ("Competency questions", "§5.1"),
    "prohibited_use": ("Prohibited uses", "§5.8"),
    "gold_concept": ("Gold concepts", "§5.3"),
    "gold_entity": ("Gold entities", "§5.2"),
    "gold_relationship": ("Gold relationships", "§5.4"),
    "gold_search_question": ("Search questions", "§5.5"),
    "gold_retrieval_question": ("AI retrieval questions", "§5.6"),
    "reasoning_expectation": ("Reasoning expectations", "§5.7"),
}

#: The order the pack reads in — the guide's suggested order of work (§10),
#: which is also the order in which each type makes the next one cheaper.
PACK_ORDER: tuple[str, ...] = (
    "competency_question",
    "prohibited_use",
    "gold_concept",
    "gold_entity",
    "gold_relationship",
    "gold_search_question",
    "gold_retrieval_question",
    "reasoning_expectation",
)

_CONTEXT = 220  # characters of passage either side of a span


# ---------------------------------------------------------------------------
# The Markdown review pack
# ---------------------------------------------------------------------------


def _quote(text: str, width: int = 92) -> str:
    return "\n".join(f"> {line}" for line in textwrap.wrap(text, width)) or "> —"


def _excerpt(resolution: Resolution) -> str:
    """The passage around the span, with the span itself marked."""
    text = resolution.passage_text
    span = resolution.record.get("span")
    if not text:
        return ""
    if not span:
        return _quote(text[: _CONTEXT * 2] + ("…" if len(text) > _CONTEXT * 2 else ""))
    start, end = span
    left = max(0, start - _CONTEXT)
    right = min(len(text), end + _CONTEXT)
    marked = (
        ("…" if left else "")
        + text[left:start]
        + "**"
        + text[start:end]
        + "**"
        + text[end:right]
        + ("…" if right < len(text) else "")
    )
    return _quote(marked)


def _field_lines(record: dict[str, Any], skip: tuple[str, ...] = ()) -> list[str]:
    lines: list[str] = []
    for key, value in record.items():
        if key in skip or key in ("id", "approved_by", "approved_date"):
            continue
        if value is None or value == [] or value == "":
            rendered = "—"
        elif isinstance(value, bool):
            # Rendered lower-case: a reader correcting a cell should see the
            # value the schema uses, not Python's spelling of it.
            rendered = "true" if value else "false"
        elif key == "span" and isinstance(value, list) and len(value) == 2:
            rendered = f"[{value[0]}, {value[1]}]"
        elif isinstance(value, list):
            rendered = (
                "; ".join(
                    ", ".join(f"{k}={v}" for k, v in item.items())
                    if isinstance(item, dict)
                    else str(item)
                    for item in value
                )
                or "—"
            )
        elif isinstance(value, dict):
            rendered = "; ".join(
                f"{k}: {', '.join(map(str, v)) if isinstance(v, list) else v}"
                for k, v in value.items()
            )
        else:
            rendered = str(value).strip().replace("\n", " ")
        lines.append(f"| `{key}` | {rendered} |")
    return lines


def render_pack(
    seed: SeedSet,
    resolutions: tuple[Resolution, ...],
    corpus: Corpus | None = None,
    *,
    generated: str | None = None,
) -> str:
    """The whole seed set as one readable document."""
    stamp = generated or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    by_type: dict[str, list[Resolution]] = {}
    for resolution in resolutions:
        by_type.setdefault(resolution.envelope.record_type, []).append(resolution)

    out: list[str] = [
        "<!-- Generated by tm_knowledge.stage0.seedpack. Do not hand-edit. -->",
        "",
        "# Seed review pack — s 43 examples, for correction",
        "",
        "> **Nothing in this document is approved, and none of it is project",
        "> content.** Every record below was written by a machine to show what a",
        "> Stage 0 record *looks like* when it is filled in over the pilot area.",
        "> The legal judgements in it are unverified and some of them are wrong.",
        "> That is the design: it is easier to correct a wrong answer than to",
        "> compose a right one from an empty form, and your corrections are what",
        "> becomes the real gold set (ADR-0043).",
        "",
        "> **What to do with it.** Read a record, read the passage under it, and",
        "> say one of three things: *correct*, *amend* (and say how), or *reject*",
        "> (and say why). A rejection is as useful as a correction — it tells us",
        "> the shape was wrong, not just the wording. You do not have to finish.",
        "",
    ]

    rows = [
        "| | |",
        "|---|---|",
        f"| Records in this pack | {len(resolutions)} |",
    ]
    for record_type in PACK_ORDER:
        if by_type.get(record_type):
            label = PACK_HEADINGS[record_type][0]
            rows.append(f"| — {label} | {len(by_type[record_type])} |")
    if corpus is not None:
        rows.append(
            f"| Pinned snapshot | `{corpus.pin.repo}` @ `{corpus.pin.commit[:12]}` |"
        )
    rows.append(f"| Generated | {stamp} |")
    out.extend(rows)
    out.append("")

    out.extend(
        [
            "## How a record is laid out",
            "",
            "Each one carries a **seed id** (`SEED-…`) — quote that when you write",
            "a correction, because the record's own id may change and the seed id",
            "will not. **Why this example** says what the record is here to",
            "demonstrate; if the record is wrong but the point is right, amend it.",
            "Where a record rests on a passage, the passage is quoted underneath",
            "with the exact span in **bold**.",
            "",
            "---",
            "",
        ]
    )

    for record_type in PACK_ORDER:
        group = by_type.get(record_type)
        if not group:
            continue
        label, section = PACK_HEADINGS[record_type]
        out.append(f"## {label} ({len(group)}) — guide {section}")
        out.append("")
        for resolution in group:
            out.extend(_render_record(resolution))
        out.append("")

    return "\n".join(out).rstrip() + "\n"


def _render_record(resolution: Resolution) -> list[str]:
    envelope = resolution.envelope
    record = resolution.record
    out = [
        f"### `{envelope.seed_id}` → `{record.get('id')}`",
        "",
        f"**Why this example.** {envelope.why_this_example.strip()}",
        "",
    ]
    if resolution.heading_path:
        out.append(f"**Where.** {' › '.join(resolution.heading_path)}")
        out.append("")
    out.append("| field | value |")
    out.append("|---|---|")
    out.extend(_field_lines(record))
    out.append("")
    excerpt = _excerpt(resolution)
    if excerpt:
        out.append(f"**Passage** — `{record.get('source_ref')}`")
        out.append("")
        out.append(excerpt)
        out.append("")
    out.append(
        "**Verdict:** ☐ correct  ☐ amend → ______________________  ☐ reject → why?"
    )
    out.append("")
    out.append("---")
    out.append("")
    return out


# ---------------------------------------------------------------------------
# The review workbook
# ---------------------------------------------------------------------------


def _review_guide() -> list[tuple[str, str]]:
    return [
        (
            "This workbook is a draft, not a form",
            "Every row is a machine-written example over s 43. None of it is "
            "approved and some of it is wrong. Correct it in place.",
        ),
        (
            "Three columns at the right-hand end are yours",
            "`verdict` is one of: " + ", ".join(VERDICTS) + ". `correction` is "
            "free text — say what is wrong in your own words rather than trying "
            "to write the field. `seed_id` is the record's handle in the review "
            "pack; leave it alone.",
        ),
        (
            "Correcting a cell is the same as writing it",
            "Type over any value you disagree with. There is nothing special "
            "about a value being pre-filled — it carries no authority at all "
            "until your name is against it.",
        ),
        (
            "Delete rows that should not exist",
            "A record that is the wrong shape, not just the wrong words, should "
            "be deleted or marked `reject`. That is a finding, not a failure.",
        ),
        (
            "Add rows freely",
            "An empty row at the bottom of a sheet is a new record. Allocate the "
            "next id in the series; do not reuse a deleted one.",
        ),
        (
            "Spans and hashes fill themselves",
            "Leave `span.start`, `span.end` and `source_content_hash` alone. If "
            "you change a `surface` or a `supporting_text`, they are recomputed "
            "from the snapshot — you never type an offset.",
        ),
        (
            "approved_by and approved_date are what make a row real",
            "A row without them is transcribed and reported as awaiting "
            "approval. A row with them is approved knowledge, and downstream "
            "work will rely on it.",
        ),
        (
            "Nothing is transcribed until you hand it back",
            "`tmk-transcribe <file> --write` reads this layout into eval/gold/. "
            "It reshapes and never supplies: a blank judgement stays blank and "
            "is reported.",
        ),
    ]


def build_workbook(seed: SeedSet, resolutions: tuple[Resolution, ...], *,
                   generated: str | None = None):
    """The intake layout, pre-filled, plus the review columns."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    records: dict[str, list[dict[str, Any]]] = {
        record_type: [] for record_type in RECORD_TYPES
    }
    seeds_for: dict[str, list[str]] = {record_type: [] for record_type in RECORD_TYPES}
    for resolution in resolutions:
        record_type = resolution.envelope.record_type
        records.setdefault(record_type, []).append(resolution.record)
        seeds_for.setdefault(record_type, []).append(resolution.envelope.seed_id)

    book = workbook_module.build(generated=generated)
    workbook_module.fill(book, records)

    for spec in sheets():
        sheet = book[spec.name]
        first = len(spec.columns) + 1
        rows = sheet.max_row
        for offset, header in enumerate(REVIEW_COLUMNS):
            index = first + offset
            cell = sheet.cell(row=1, column=index, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFFCE4D6")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = 20 if header != "correction" else 60
            if header == "verdict":
                cell.comment = Comment(
                    "correct · amend · reject. Leave blank if you have not read "
                    "this row yet.",
                    "tmk-seed",
                    height=120,
                    width=300,
                )
                validation = DataValidation(
                    type="list",
                    formula1='"' + ",".join(VERDICTS) + '"',
                    allow_blank=True,
                )
                sheet.add_data_validation(validation)
                validation.add(f"{letter}2:{letter}{max(rows, 2) + 200}")
            if header == "correction":
                cell.comment = Comment(
                    "What is wrong, in your own words. You do not have to write "
                    "the corrected field — saying what the record gets wrong is "
                    "the part only you can do.",
                    "tmk-seed",
                    height=140,
                    width=340,
                )

        if spec.is_child:
            continue
        ids = seeds_for.get(spec.record_type, [])
        column = first + REVIEW_COLUMNS.index("seed_id")
        for offset, seed_id in enumerate(ids):
            sheet.cell(row=2 + offset, column=column, value=seed_id)

    _rewrite_guide(book, seed, generated=generated)
    book.properties.title = "TM-Knowledge — seed review (s 43)"
    book.properties.description = (
        "Machine-written example records over s 43, for expert correction. "
        "Not approved, not project content. Generated by tmk-seed on "
        f"{generated or datetime.now(timezone.utc).strftime('%Y-%m-%d')}."
    )
    return book


def _rewrite_guide(book, seed: SeedSet, *, generated: str | None = None) -> None:
    from openpyxl.styles import Alignment, Font

    name = workbook_module.GUIDE_SHEET
    if name in book.sheetnames:
        book.remove(book[name])
    sheet = book.create_sheet(name, 0)
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 96

    sheet["A1"] = "Stage 0 seed review — s 43"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = (
        f"{seed.total} machine-written example records, for correction. Nothing "
        "here is approved and none of it is project content. It exists because "
        "correcting a wrong answer is easier than composing a right one from a "
        "blank form — your corrections are what becomes the gold set."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A2:B2")
    sheet.row_dimensions[2].height = 48

    row = 4
    for heading, body in _review_guide():
        sheet.cell(row=row, column=1, value=heading).font = Font(bold=True)
        cell = sheet.cell(row=row, column=2, value=body)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 46
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Where the readable version is").font = Font(
        bold=True
    )
    sheet.cell(
        row=row,
        column=2,
        value=(
            "data/derived/seed-review-pack.md prints every record with the "
            "passage it rests on quoted underneath. Read there, correct here — "
            "or ignore this file entirely and mark up the pack."
        ),
    ).alignment = Alignment(wrap_text=True, vertical="top")


def write_workbook(
    path: Path,
    seed: SeedSet,
    resolutions: tuple[Resolution, ...],
    *,
    generated: str | None = None,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(seed, resolutions, generated=generated).save(path)
    return path
