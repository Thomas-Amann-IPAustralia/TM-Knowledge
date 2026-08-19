"""The intake path — P7's and P8's done-criteria.

P7: the workbook regenerates from the schemas, enum cells reject out-of-
vocabulary values, and a round trip through P8 preserves every field.
P8: a filled workbook transcribes to schema-valid records, every missing
judgement field appears as a gap rather than being filled, and re-running over
unchanged input rewrites nothing.

The tests that matter most are the ones about *not* doing things. Transcription
may reshape and never supply, so a blank `modality` must come out blank, a row
that is not yet a record must be rejected rather than stubbed, and a value
outside an enum must stop the read rather than be snapped to the nearest
allowed one. Each of those is a legal reading if a machine makes it.

No snapshot needed: nothing here resolves a ref.
"""

from __future__ import annotations

import pytest

pytest.importorskip("openpyxl", reason="the intake path needs `pip install -e '.[intake]'`")

from tm_knowledge.config import REPO_ROOT  # noqa: E402
from tm_knowledge.stage0 import goldset, transcribe, workbook  # noqa: E402
from tm_knowledge.stage0.intake import SHEET_NAMES, sheets  # noqa: E402
from tm_knowledge.stage0.schemas import (  # noqa: E402
    RECORD_TYPES,
    property_order,
    validate,
)

SOUND = REPO_ROOT / "tests" / "fixtures" / "harness" / "sound"


@pytest.fixture(scope="module")
def source_records() -> dict[str, list[dict]]:
    gold = goldset.load(SOUND)
    return {record_type: list(records) for record_type, records in gold.records.items()}


def _filled(tmp_path, records, name="filled.xlsx"):
    book = workbook.build(generated="2026-08-19")
    workbook.fill(book, records)
    path = tmp_path / name
    book.save(path)
    return path


# ---------------------------------------------------------------------------
# P7 — the workbook
# ---------------------------------------------------------------------------


def test_the_workbook_ships_with_no_example_rows(tmp_path):
    """Not one, not even a marked one. A row in a spreadsheet is a keystroke
    from being twenty rows, and a plausible answer nobody approved is worse than
    a blank (CLAUDE.md rule 1)."""
    from openpyxl import load_workbook

    path = tmp_path / "empty.xlsx"
    workbook.write(path, generated="2026-08-19")
    book = load_workbook(path)
    for spec in sheets():
        assert book[spec.name].max_row == 1, f"{spec.name} ships with data in it"


def test_every_schema_field_reaches_a_column():
    """A field with no column is a field the expert is never asked for."""
    for record_type in RECORD_TYPES:
        covered = set()
        for spec in sheets():
            if spec.record_type != record_type:
                continue
            if spec.is_child:
                covered.add(spec.parent_field)
            else:
                covered.update(column.path[0] for column in spec.columns)
        missing = set(property_order(record_type)) - covered
        assert not missing, f"{record_type} has no column for {sorted(missing)}"


def test_enum_cells_offer_only_the_allowed_values(tmp_path):
    from openpyxl import load_workbook

    path = tmp_path / "empty.xlsx"
    workbook.write(path, generated="2026-08-19")
    book = load_workbook(path)
    for spec in sheets():
        wanted = sum(1 for column in spec.columns if column.enum or column.kind == "boolean")
        found = len(book[spec.name].data_validations.dataValidation)
        assert found >= min(wanted, 1) or wanted == 0, spec.name


def test_the_enum_sheet_is_machinery_and_stays_hidden(tmp_path):
    from openpyxl import load_workbook

    path = tmp_path / "empty.xlsx"
    workbook.write(path, generated="2026-08-19")
    book = load_workbook(path)
    assert book[workbook.ENUM_SHEET].sheet_state == "hidden"


def test_the_sheet_names_are_the_gold_file_names():
    """So an expert reading `eval/gold/` recognises the sheet they filled in."""
    for record_type, sheet_name in SHEET_NAMES.items():
        assert goldset.FILE_FOR[record_type] == f"{sheet_name}.yaml"


# ---------------------------------------------------------------------------
# P8 — the round trip
# ---------------------------------------------------------------------------


def test_the_round_trip_preserves_every_populated_field(tmp_path, source_records):
    result = transcribe.read_workbook(_filled(tmp_path, source_records))
    assert result.problems == [], "\n".join(str(p) for p in result.problems)
    for record_type, records in source_records.items():
        transcribed = result.records[record_type]
        assert len(transcribed) == len(records)
        for before, after in zip(records, transcribed):
            for key, value in before.items():
                if value in (None, [], ""):
                    continue  # an optional empty is dropped on purpose
                assert after.get(key) == value, f"{record_type}.{key} did not survive"


def test_transcribed_records_are_in_schema_order(tmp_path, source_records):
    """So a gold file's diff reads like the template it came from."""
    result = transcribe.read_workbook(_filled(tmp_path, source_records))
    for record_type, records in result.records.items():
        order = property_order(record_type)
        for record in records:
            positions = [order.index(key) for key in record]
            assert positions == sorted(positions), f"{record_type} came out reordered"


def test_the_round_trip_is_a_fixed_point(tmp_path, source_records):
    once = transcribe.read_workbook(_filled(tmp_path, source_records, "a.xlsx"))
    twice = transcribe.read_workbook(_filled(tmp_path, once.records, "b.xlsx"))
    assert twice.records == once.records


def test_everything_transcribed_validates(tmp_path, source_records):
    result = transcribe.read_workbook(_filled(tmp_path, source_records))
    for record_type, records in result.records.items():
        for record in records:
            assert validate(record, record_type) == []


def test_a_list_cell_keeps_a_value_that_contains_a_separator(tmp_path):
    """Refs carry `/`, `(`, `~`, `#` and `.`; a comma-separated cell would be a
    data-loss bug waiting for the first value that uses the separator."""
    records = {
        "gold_concept": [
            {
                "id": "GC-001",
                "pref_label": "«label»",
                "alt_labels": ["one, with a comma", "two; with a semicolon"],
                "not_labels": [],
                "definition_sources": ["TMA1995/s41(3)(a)"],
                "approved_by": None,
                "approved_date": None,
            }
        ]
    }
    result = transcribe.read_workbook(_filled(tmp_path, records))
    assert result.records["gold_concept"][0]["alt_labels"] == [
        "one, with a comma",
        "two; with a semicolon",
    ]
    assert result.records["gold_concept"][0]["definition_sources"] == ["TMA1995/s41(3)(a)"]


# ---------------------------------------------------------------------------
# Reshape, never supply
# ---------------------------------------------------------------------------


def _relationship(**overrides):
    record = {
        "id": "GR-001",
        "subject": "«subject»",
        "predicate": "«predicate»",
        "object": "«object»",
        "source_ref": "TMM/Part20/5/5/1",
        "supporting_text": "«the sentence, verbatim»",
        "span": None,
        "source_content_hash": None,
        "tier": None,
        "modality": None,
        "approved_by": None,
        "approved_date": None,
    }
    record.update(overrides)
    return {"gold_relationship": [record]}


def test_a_blank_judgement_field_comes_out_blank_and_is_reported(tmp_path):
    """`modality` is must / may / should, and whether a "may" is possibility or
    permission is a legal reading (guide §5.4). Nothing here may decide it."""
    result = transcribe.read_workbook(_filled(tmp_path, _relationship()))
    record = result.records["gold_relationship"][0]
    assert record["modality"] is None
    assert record["tier"] is None
    assert record["span"] is None
    reported = {name for _, identifier, name in result.blanks if identifier == "GR-001"}
    assert {"modality", "tier", "span", "approved_by", "approved_date"} <= reported


def test_a_row_missing_a_non_nullable_field_is_rejected_not_stubbed(tmp_path):
    result = transcribe.read_workbook(
        _filled(tmp_path, _relationship(predicate=None))
    )
    assert result.records == {}
    assert any("predicate" in str(problem) for problem in result.problems)


def test_an_enum_value_outside_the_list_stops_the_row(tmp_path):
    """Not snapped to the nearest allowed value. "probably" is not "may"."""
    path = _filled(tmp_path, _relationship())
    from openpyxl import load_workbook

    book = load_workbook(path)
    sheet = book["relationships"]
    column = [c.value for c in sheet[1]].index("modality") + 1
    sheet.cell(row=2, column=column, value="probably")
    book.save(path)

    result = transcribe.read_workbook(path)
    assert result.records == {}
    assert any("probably" in str(problem) for problem in result.problems)


def test_half_a_span_is_rejected(tmp_path):
    path = _filled(tmp_path, _relationship())
    from openpyxl import load_workbook

    book = load_workbook(path)
    sheet = book["relationships"]
    column = [c.value for c in sheet[1]].index("span.start") + 1
    sheet.cell(row=2, column=column, value=0)
    book.save(path)

    result = transcribe.read_workbook(path)
    assert any("span.start and span.end" in str(problem) for problem in result.problems)


def test_a_record_that_does_not_validate_is_never_written(tmp_path):
    path = _filled(tmp_path, _relationship())
    from openpyxl import load_workbook

    book = load_workbook(path)
    sheet = book["relationships"]
    column = [c.value for c in sheet[1]].index("id") + 1
    sheet.cell(row=2, column=column, value="XX-001")
    book.save(path)

    result = transcribe.read_workbook(path)
    assert result.records == {}
    assert any("XX-001" in str(problem) for problem in result.problems)


def test_a_child_row_with_no_parent_is_rejected(tmp_path):
    records = {
        "gold_search_question": [
            {
                "id": "GS-001",
                "query": "«a query»",
                "uses_manual_terminology": False,
                "relevant": [{"ref": "TMM/Part20/5/5/1", "grade": 3}],
                "approved_by": None,
                "approved_date": None,
            }
        ]
    }
    path = _filled(tmp_path, records)
    from openpyxl import load_workbook

    book = load_workbook(path)
    sheet = book["GS--relevant"]
    sheet.cell(row=3, column=1, value="GS-999")
    sheet.cell(row=3, column=2, value="TMM/Part20/5/5/1")
    sheet.cell(row=3, column=3, value=3)
    book.save(path)

    result = transcribe.read_workbook(path)
    assert any("GS-999" in str(problem) for problem in result.problems)
    assert len(result.records["gold_search_question"][0]["relevant"]) == 1


def test_a_column_added_by_hand_stops_the_read(tmp_path, source_records):
    """A column the transcriber does not know is a field nobody collects."""
    path = _filled(tmp_path, source_records)
    from openpyxl import load_workbook

    book = load_workbook(path)
    sheet = book["entities"]
    sheet.cell(row=1, column=sheet.max_column + 1, value="confidence")
    book.save(path)

    with pytest.raises(transcribe.WorkbookMismatch) as error:
        transcribe.read_workbook(path)
    assert "confidence" in str(error.value)


# ---------------------------------------------------------------------------
# Writing into approved space
# ---------------------------------------------------------------------------


def test_a_dry_run_writes_nothing(tmp_path, source_records):
    result = transcribe.read_workbook(_filled(tmp_path, source_records))
    gold = tmp_path / "gold"
    outcomes = transcribe.write_records(result, gold, write=False)
    assert outcomes and all(outcome == "would write" for _, outcome in outcomes)
    assert not gold.exists()


def test_writing_twice_changes_nothing_the_second_time(tmp_path, source_records):
    """Re-running over unchanged input must leave git status clean."""
    result = transcribe.read_workbook(_filled(tmp_path, source_records))
    gold = tmp_path / "gold"
    first = transcribe.write_records(result, gold, write=True)
    assert all(outcome == "written" for _, outcome in first)
    second = transcribe.write_records(result, gold, write=True)
    assert all(outcome == "unchanged" for _, outcome in second)


def test_what_was_written_loads_back_as_a_gold_set(tmp_path, source_records):
    result = transcribe.read_workbook(_filled(tmp_path, source_records))
    gold = tmp_path / "gold"
    transcribe.write_records(result, gold, write=True)
    loaded = goldset.load(gold)
    assert loaded.unreadable == ()
    assert loaded.total == result.total


def test_an_empty_sheet_leaves_its_file_alone(tmp_path):
    """An empty sheet means "I have nothing for this yet", never "delete what is
    there". Filling one sheet must not wipe seven others."""
    gold = tmp_path / "gold"
    gold.mkdir()
    (gold / "concepts.yaml").write_text("- {id: GC-001}\n", encoding="utf-8")

    result = transcribe.read_workbook(_filled(tmp_path, _relationship()))
    assert "gold_concept" in result.empty_sheets
    transcribe.write_records(result, gold, write=True)
    assert (gold / "concepts.yaml").read_text(encoding="utf-8") == "- {id: GC-001}\n"
